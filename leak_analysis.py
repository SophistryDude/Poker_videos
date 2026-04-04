"""
Leak Analysis — Why are players winning or losing?

Classifies 100K synthetic players by their primary leak/strength.
A "leak" is the main thing holding back their EV.
A "strength" is what's driving their profit.

Leaks:
  - Too Aggressive: high agg, low CR, busts before cashing enough
  - Too Passive: low agg, high CR but small stacks, min-cash grinder
  - Can't Close: makes FTs but terrible top-3 conversion (bubble boy)
  - No Final Tables: cashes but never goes deep (flat distribution)
  - Min-Cash Machine: high CR, almost all cashes are 10th+
  - Small Sample: <100 tournaments, stats unreliable
  - Skill Ceiling: balanced style but just not good enough

Strengths (for winners):
  - Optimal Aggression: aggression near 0.55-0.70, good CR/FT balance
  - Deep Run Specialist: high FT rate, stacks convert to final tables
  - Closer: high top-3 conversion, wins when at FT
  - Volume Edge: large sample, consistent small edge compounding
  - Complete Player: strong across all metrics
"""

import csv
import time
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict

from stake_move_model import (
    TIERS, TIER_ORDER, project_cash_rate, project_ft_rate,
    project_placements, calc_ev, apply_aggression,
)

HRS_PER_ENTRY = 6.5

PLAYER_TIERS = [
    "Brand New", "Bad Reg", "Average Player", "Slightly Profitable",
    "Good Reg", "Great Reg", "Low Level Pro", "Mid Level Pro",
    "High Level Pro", "Best in the World",
]


def get_evs(p):
    """Calculate EVs at all tiers for a player."""
    t = p["tournaments"]
    agg = p.get("aggression", 0.5)
    if t == 0:
        return None

    current_placements = {
        "1st": p["1st"], "2nd": p["2nd"], "3rd": p["3rd"],
        "4th-9th": p["4th-9th"], "10th+": p["10th+"],
    }

    evs = {}
    for i, tier_key in enumerate(TIER_ORDER):
        tier = TIERS[tier_key]
        if i == 0:
            placements = current_placements
        else:
            cr = project_cash_rate(p["cash_rate"], i)
            ft = project_ft_rate(p["ft_rate"], p["cash_rate"], cr)
            placements = project_placements(p, cr, ft, aggression=agg)
        evs[tier_key] = calc_ev(placements, t, tier["payouts"], tier["buyin"])

    return evs


def classify_leak(p, evs):
    """Classify a player's primary leak or strength."""
    cr = p["cash_rate"]
    ft = p["ft_rate"]
    t3 = p["top3_conv"]
    agg = p.get("aggression", 0.5)
    t = p["tournaments"]
    cashes = p["total_cashes"]
    tenth_plus = p["10th+"]

    ev_low = evs["low"]
    best_ev = max(evs.values())
    best_tier = max(evs, key=lambda k: evs[k])

    ft_to_cash = ft / cr if cr > 0 else 0
    mincash_pct = tenth_plus / max(1, cashes)

    # Classify losing players (negative EV at their current level)
    if ev_low <= 0:
        # Too Aggressive: high aggression is cratering their CR
        # Signal: agg > 0.65 AND CR is low for their tier
        if agg >= 0.65 and cr < 0.15:
            return "Leak: Too Aggressive"

        # Too Passive: high CR but stacks are too small to go deep
        # Signal: agg < 0.35 AND decent CR but terrible FT rate
        if agg <= 0.35 and cr > 0.12 and ft_to_cash < 0.25:
            return "Leak: Too Passive"

        # Min-Cash Machine: most cashes are 10th+, never goes deep
        if mincash_pct > 0.80 and cashes >= 5:
            return "Leak: Min-Cash Machine"

        # Can't Close: makes FTs but can't convert top-3
        if ft_to_cash > 0.35 and t3 < 0.25 and ft > 0.03:
            return "Leak: Can't Close"

        # No Final Tables: cashes but almost never FTs
        if cr > 0.10 and ft_to_cash < 0.20 and ft < 0.03:
            return "Leak: No Final Tables"

        # Small Sample: could just be variance
        if t < 100:
            return "Leak: Small Sample"

        # Generic — just not skilled enough
        return "Leak: Skill Ceiling"

    # Classify winning players by what drives their profit
    else:
        # Small Sample winner: might be running hot
        if t < 100:
            return "Strength: Small Sample (caution)"

        # Complete Player: strong across all metrics
        if cr > 0.28 and ft_to_cash > 0.30 and t3 > 0.40 and agg >= 0.40 and agg <= 0.70:
            return "Strength: Complete Player"

        # Optimal Aggression: near the sweet spot
        if agg >= 0.55 and agg <= 0.75 and cr > 0.15 and ft_to_cash > 0.30:
            return "Strength: Optimal Aggression"

        # Deep Run Specialist: high FT rate relative to cashes
        if ft_to_cash > 0.45 and ft > 0.08:
            return "Strength: Deep Run Specialist"

        # Closer: high top-3 conversion at FT
        if t3 > 0.55 and ft > 0.05:
            return "Strength: Closer"

        # Volume Edge: moderate stats but big sample compounds
        if t > 1000 and cr > 0.15 and ev_low > 0 and ev_low < 200:
            return "Strength: Volume Edge"

        # Passive but profitable: high CR compensates for small stacks
        if agg < 0.35 and cr > 0.22:
            return "Strength: Passive Grinder"

        # Aggressive and profitable: low CR but big scores
        if agg > 0.65 and cr < 0.22 and best_ev > 500:
            return "Strength: Aggressive + Skilled"

        return "Strength: Solid Overall"


def load_players(csv_path):
    players = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            players.append({
                "id": int(row["id"]),
                "tier": row["tier"],
                "tournaments": int(row["tournaments"]),
                "1st": int(row["1st"]),
                "2nd": int(row["2nd"]),
                "3rd": int(row["3rd"]),
                "4th-9th": int(row["4th-9th"]),
                "10th+": int(row["10th+"]),
                "total_cashes": int(row["total_cashes"]),
                "cash_rate": float(row["cash_rate"]),
                "ft_rate": float(row["ft_rate"]),
                "top3_conv": float(row["top3_conv"]),
                "aggression": float(row.get("aggression", 0.5)),
            })
    return players


def build_workbook(results, leak_details):
    wb = openpyxl.Workbook()

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    halign = Alignment(horizontal="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    leak_colors = {
        "Leak: Too Aggressive": "e63946",
        "Leak: Too Passive": "ff6b6b",
        "Leak: Min-Cash Machine": "d62828",
        "Leak: Can't Close": "f77f00",
        "Leak: No Final Tables": "e76f51",
        "Leak: Small Sample": "adb5bd",
        "Leak: Skill Ceiling": "6c757d",
        "Strength: Complete Player": "06d6a0",
        "Strength: Optimal Aggression": "1b9e77",
        "Strength: Deep Run Specialist": "118ab2",
        "Strength: Closer": "073b4c",
        "Strength: Volume Edge": "7209b7",
        "Strength: Passive Grinder": "48bfe3",
        "Strength: Aggressive + Skilled": "2d6a4f",
        "Strength: Solid Overall": "52b788",
        "Strength: Small Sample (caution)": "ced4da",
    }

    # ── Sheet 1: Leak Summary ──
    ws1 = wb.active
    ws1.title = "Leak Summary"

    s1_headers = ["Classification", "Count", "%", "Avg CR", "Avg Agg", "Avg FT Rate",
                  "Avg T3 Conv", "Avg EV@$300", "Avg Tourns", "Min-Cash %"]
    for col, h in enumerate(s1_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    # Sort: leaks first, then strengths
    all_classes = sorted(set(r["classification"] for r in results),
                         key=lambda x: (0 if "Leak" in x else 1, x))

    for row_idx, cls in enumerate(all_classes, 2):
        recs = [r for r in results if r["classification"] == cls]
        n = len(recs)
        color = leak_colors.get(cls, "FFFFFF")

        vals = [
            cls, n, round(n / len(results), 4),
            round(np.mean([r["cash_rate"] for r in recs]), 4),
            round(np.mean([r["aggression"] for r in recs]), 3),
            round(np.mean([r["ft_rate"] for r in recs]), 4),
            round(np.mean([r["top3_conv"] for r in recs]), 4),
            round(np.mean([r["ev_low"] for r in recs]), 0),
            round(np.mean([r["tournaments"] for r in recs])),
            round(np.mean([r["10th+"] / max(1, r["total_cashes"]) for r in recs]), 4),
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if col_idx == 1:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            elif col_idx == 3:
                cell.number_format = "0.0%"
            elif col_idx in (4, 6, 7, 10):
                cell.number_format = "0.0%"
            elif col_idx == 5:
                cell.number_format = "0.00"
            elif col_idx == 8:
                cell.number_format = '$#,##0'

    for col in range(1, len(s1_headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 18
    ws1.column_dimensions["A"].width = 32
    ws1.freeze_panes = "B2"

    # ── Sheet 2: Leaks by Player Tier ──
    ws2 = wb.create_sheet("Leaks by Player Tier")

    tier_headers = ["Player Tier"] + all_classes
    for col, h in enumerate(tier_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", wrap_text=True, text_rotation=45)
        cell.border = border

    tier_colors = {
        "Brand New": "FFE0E0", "Bad Reg": "FFD0C0", "Average Player": "FFE8B0",
        "Slightly Profitable": "FFFFC0", "Good Reg": "D8F0D8", "Great Reg": "B0E0B0",
        "Low Level Pro": "B0D8F0", "Mid Level Pro": "90C0E8",
        "High Level Pro": "C0B0F0", "Best in the World": "FFD700",
    }

    for row_idx, tn in enumerate(PLAYER_TIERS, 2):
        tr = [r for r in results if r["tier"] == tn]
        counts = Counter(r["classification"] for r in tr)
        n = len(tr)

        fill = PatternFill(start_color=tier_colors.get(tn, "FFFFFF"),
                          end_color=tier_colors.get(tn, "FFFFFF"), fill_type="solid")

        ws2.cell(row=row_idx, column=1, value=tn).fill = fill
        ws2.cell(row=row_idx, column=1).border = border

        for col_idx, cls in enumerate(all_classes, 2):
            count = counts.get(cls, 0)
            pct = count / max(1, n)
            cell = ws2.cell(row=row_idx, column=col_idx, value=f"{count} ({pct:.0%})")
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if pct > 0.3:
                cell.font = Font(bold=True)
            if pct > 0.2:
                color = leak_colors.get(cls, "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(tier_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 16
    ws2.column_dimensions["A"].width = 22
    ws2.freeze_panes = "B2"
    ws2.row_dimensions[1].height = 80

    # ── Sheet 3: Aggression vs EV scatter data ──
    ws3 = wb.create_sheet("Aggression vs EV")

    ws3.cell(row=1, column=1, value="Aggression vs EV by Classification").font = Font(bold=True, size=13)

    agg_headers = ["Classification", "Avg Agg", "Avg EV@$300", "Avg EV@$600",
                   "Avg EV@$1K", "Avg EV@$1.8K", "Optimal Tier",
                   "P25 Agg", "P75 Agg", "EV Spread (P75-P25)"]
    for col, h in enumerate(agg_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    for row_idx, cls in enumerate(all_classes, 4):
        recs = [r for r in results if r["classification"] == cls]
        aggs = [r["aggression"] for r in recs]
        color = leak_colors.get(cls, "FFFFFF")

        # Get EVs at all tiers
        ev_lows = [r["ev_low"] for r in recs]
        ev_mls = [r["ev_mid_low"] for r in recs]
        ev_mids = [r["ev_mid"] for r in recs]
        ev_highs = [r["ev_high"] for r in recs]

        # Find most common best tier
        best_tiers = Counter()
        for r in recs:
            evs = {"$300": r["ev_low"], "$600": r["ev_mid_low"],
                   "$1K": r["ev_mid"], "$1.8K": r["ev_high"]}
            best = max(evs, key=lambda k: evs[k])
            best_tiers[best] += 1
        top_tier = best_tiers.most_common(1)[0][0] if best_tiers else "?"

        vals = [
            cls, round(np.mean(aggs), 3),
            round(np.mean(ev_lows), 0), round(np.mean(ev_mls), 0),
            round(np.mean(ev_mids), 0), round(np.mean(ev_highs), 0),
            top_tier,
            round(np.percentile(aggs, 25), 3), round(np.percentile(aggs, 75), 3),
            round(np.percentile(ev_lows, 75) - np.percentile(ev_lows, 25), 0),
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if col_idx == 1:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            elif col_idx == 2 or col_idx in (8, 9):
                cell.number_format = "0.00"
            elif col_idx in (3, 4, 5, 6, 10):
                cell.number_format = '$#,##0'

    for col in range(1, len(agg_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 18
    ws3.column_dimensions["A"].width = 32

    # ── Sheet 4: Sample players from each leak category ──
    ws4 = wb.create_sheet("Examples (20 per leak)")

    s4_headers = ["Classification", "Tier", "Tourns", "Agg", "CR", "FT Rate",
                  "T3 Conv", "MinCash%", "EV@$300", "EV@$600", "EV@$1K", "EV@$1.8K"]
    for col, h in enumerate(s4_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    row_idx = 2
    for cls in all_classes:
        recs = [r for r in results if r["classification"] == cls]
        # Sample 20, or all if fewer
        sample_idx = np.random.choice(len(recs), size=min(20, len(recs)), replace=False)
        sample = [recs[i] for i in sample_idx]
        color = leak_colors.get(cls, "FFFFFF")

        for r in sample:
            mc_pct = r["10th+"] / max(1, r["total_cashes"])
            vals = [cls, r["tier"], r["tournaments"], r["aggression"],
                    r["cash_rate"], r["ft_rate"], r["top3_conv"],
                    round(mc_pct, 3), r["ev_low"], r["ev_mid_low"],
                    r["ev_mid"], r["ev_high"]]

            for col_idx, val in enumerate(vals, 1):
                cell = ws4.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
                if col_idx == 1:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF", size=9)
                elif col_idx in (5, 6, 7, 8):
                    cell.number_format = "0.0%"
                elif col_idx == 4:
                    cell.number_format = "0.00"
                elif col_idx >= 9:
                    cell.number_format = '$#,##0'
            row_idx += 1

    for col in range(1, len(s4_headers) + 1):
        ws4.column_dimensions[get_column_letter(col)].width = 15
    ws4.column_dimensions["A"].width = 32
    ws4.freeze_panes = "B2"

    return wb


def main():
    start = time.time()

    print("Loading 100K synthetic players...")
    players = load_players("c:/Users/nicho/OneDrive/Desktop/code/Poker_videos/synthetic_players_100k.csv")
    print(f"Loaded {len(players):,} players\n")

    print("Calculating EVs and classifying leaks...")
    results = []
    for p in players:
        evs = get_evs(p)
        if not evs:
            continue

        cls = classify_leak(p, evs)
        results.append({
            **p,
            "ev_low": round(evs["low"], 0),
            "ev_mid_low": round(evs["mid_low"], 0),
            "ev_mid": round(evs["mid"], 0),
            "ev_high": round(evs["high"], 0),
            "classification": cls,
        })

    print(f"Classified {len(results):,} players in {time.time()-start:.1f}s\n")

    # Summary
    all_classes = sorted(set(r["classification"] for r in results),
                         key=lambda x: (0 if "Leak" in x else 1, x))

    print(f"{'Classification':>35s} | {'Count':>6s} | {'%':>5s} | {'Avg CR':>6s} | {'Avg Agg':>7s} | {'Avg EV':>8s} | {'MC%':>5s}")
    print("-" * 90)

    for cls in all_classes:
        recs = [r for r in results if r["classification"] == cls]
        n = len(recs)
        pct = n / len(results)
        avg_cr = np.mean([r["cash_rate"] for r in recs])
        avg_agg = np.mean([r["aggression"] for r in recs])
        avg_ev = np.mean([r["ev_low"] for r in recs])
        mc_pct = np.mean([r["10th+"] / max(1, r["total_cashes"]) for r in recs])
        print(f"{cls:>35s} | {n:>6,d} | {pct:>4.1%} | {avg_cr:>5.1%} | {avg_agg:>6.2f} | ${avg_ev:>+7,.0f} | {mc_pct:>4.1%}")

    # Leak distribution by player tier
    print("\n--- Leak Distribution by Tier ---\n")
    leak_classes = [c for c in all_classes if "Leak" in c]
    strength_classes = [c for c in all_classes if "Strength" in c]

    for tn in PLAYER_TIERS:
        tr = [r for r in results if r["tier"] == tn]
        counts = Counter(r["classification"] for r in tr)
        n = len(tr)
        leaks = sum(v for k, v in counts.items() if "Leak" in k)
        print(f"  {tn:>22s} ({n:>5,d}): {leaks/n:>5.1%} losing")
        top3 = counts.most_common(3)
        for cls, cnt in top3:
            short = cls.replace("Leak: ", "L:").replace("Strength: ", "S:")
            print(f"    {short:>30s}: {cnt:>5,d} ({cnt/n:.0%})")

    # Save
    print("\nBuilding Excel...")
    wb = build_workbook(results, {})
    output_path = "c:/Users/nicho/OneDrive/Desktop/code/Poker_videos/leak_analysis_100k.xlsx"
    wb.save(output_path)
    print(f"Saved -> {output_path}")
    print(f"\nTotal time: {time.time()-start:.1f}s")


if __name__ == "__main__":
    main()
