"""
Stake Move Model — Should a player move up, stay, or move down?

Uses REAL payout data from PostgreSQL poker_tournaments database:
  - Tier 1: $0-$300     (avg $46, 138 tournaments — provisional, scraping more)
  - Tier 2: $301-$600   (avg $596, 130 tournaments with payouts)
  - Tier 3: $601-$1,500 (avg $1,036, 98 tournaments with payouts)
  - Tier 4: $1,500-$3,000 (avg $1,783, 34 tournaments with payouts)

Evaluates each player at all tiers using cash rate degradation curves
to model tougher fields at higher stakes.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# REAL PAYOUT DATA (from PostgreSQL — updated from DB queries)
# ============================================================

TIERS = {
    "low": {
        "label": "$0-$300",
        "buyin": 268,
        "payouts": {
            # REAL DATA: 1,686 tournaments, 26,562 payout records ($201-$300 range)
            # Avg field: 420 players, avg pool: $95,323
            "1st": 7594,
            "2nd": 5843,
            "3rd": 4298,
            "4th-9th": 2051,
            "10th+": 813,
        },
        "sample_size": 1686,
        "note": "Real data from 1,686 tournaments at Wynn/Venetian/Orleans/South Point",
    },
    "mid_low": {
        "label": "$301-$600",
        "buyin": 596,
        "payouts": {
            "1st": 52077,
            "2nd": 40797,
            "3rd": 29397,
            "4th-9th": 12331,
            "10th+": 1924,
        },
        "sample_size": 130,
        "note": "Solid sample — 130 tournaments",
    },
    "mid": {
        "label": "$601-$1,500",
        "buyin": 1036,
        "payouts": {
            "1st": 92011,
            "2nd": 69347,
            "3rd": 52400,
            "4th-9th": 20841,
            "10th+": 3645,
        },
        "sample_size": 98,
        "note": "Good sample — 98 tournaments",
    },
    "high": {
        "label": "$1,500-$3,000",
        "buyin": 1783,
        "payouts": {
            "1st": 172903,
            "2nd": 127200,
            "3rd": 96212,
            "4th-9th": 38937,
            "10th+": 5697,
        },
        "sample_size": 34,
        "note": "Decent sample — 34 tournaments",
    },
}

# Ordered tier keys for iteration
TIER_ORDER = ["low", "mid_low", "mid", "high"]


# ============================================================
# AGGRESSION MODEL
# ============================================================
# Pro consensus: ~18% cash rate is optimal because you take more
# variance pre-money to build a bigger stack.
#
# Key facts:
#   - Min cash = top 12.5% of field
#   - Average stack at min cash = 8x starting stack
#   - Bigger stack at min cash → way more FTs and wins
#   - Smaller stack at min cash → min-cash and bust
#
# Aggression coefficient (0.0 to 1.0):
#   0.0 = ultra-tight nit (survives to cash, tiny stack, min-cashes)
#   0.5 = balanced (tournament average)
#   1.0 = maximum aggression (busts more, but big stack when alive)
#
# Effects:
#   1. Cash rate: aggressive players cash LESS often
#   2. Stack at min cash: aggressive players have BIGGER stacks
#   3. Placement distribution: bigger stack → more FTs, more wins
#      smaller stack → more min-cashes

def apply_aggression(cash_rate, ft_rate, top3_conv, aggression):
    """Adjust a player's stats based on aggression coefficient.

    Returns (adjusted_cr, adjusted_ft_rate, adjusted_top3_conv, stack_mult).

    The key mechanic: aggression trades cash rate for stack size.
    Stack size at min-cash determines how deep you go.
    """
    # How much aggression deviates from neutral (0.5)
    delta = aggression - 0.5  # range: -0.5 (passive) to +0.5 (aggressive)

    # 1. Cash rate adjustment
    # Full aggression (1.0) drops CR by ~25%, full passive increases by ~25%
    cr_mult = 1.0 - delta * 0.5
    adj_cr = cash_rate * cr_mult

    # 2. Stack size at min-cash (relative to 8x avg)
    # Aggressive: 12-14x. Passive: 5-6x.
    stack_mult = 8.0 * (1.0 + delta * 1.5)
    # range: 8 * 0.25 = 2x (ultra passive) to 8 * 1.75 = 14x (ultra aggressive)
    # realistic range at 0.3-0.7 aggression: 6x to 10x

    # 3. FT rate as fraction of cashes — driven by stack size
    # Bigger stack = way more likely to reach FT from min-cash
    # At 8x avg stack, FT conversion from cash is "normal"
    # At 14x, you're a monster — FT rate per cash jumps significantly
    # At 5x, you're short and likely busting before FT
    stack_advantage = stack_mult / 8.0  # 1.0 = average
    # FT rate scales with stack advantage, but not linearly
    # (diminishing returns on huge stacks, sharp dropoff on small)
    ft_scale = stack_advantage ** 0.7  # sub-linear — 14x doesn't 1.75x your FT rate
    adj_ft = ft_rate * ft_scale

    # FT rate can't exceed cash rate (you can't FT if you didn't cash)
    adj_ft = min(adj_ft, adj_cr * 0.85)

    # 4. Top-3 conversion at FT — also stack-dependent
    # Big stack at FT = more likely to win. Small stack = bubble boy.
    t3_scale = stack_advantage ** 0.5  # even more diminishing
    adj_t3 = min(0.85, top3_conv * t3_scale)

    return adj_cr, adj_ft, adj_t3, stack_mult


def estimate_aggression(cash_rate, ft_rate, top3_conv, field_normalized=False):
    """Estimate a player's aggression from their stats.

    The key insight from pro consensus: optimal cash rate is ~18%.
    Players with higher CR than their skill level suggests are passive
    (they're folding into the money with small stacks).
    Players with lower CR are aggressive (busting more but bigger stacks).

    For SYNTHETIC players (field_normalized=True or no field size issues):
    We use the relationship between cash rate and skill tier as the
    primary signal, since FT-to-cash ratio is unreliable in small fields.

    For skill estimation, we use the EV-implied skill level:
    - A player who is +EV despite 15% CR is skilled + aggressive
    - A player who is barely +EV with 30% CR is less skilled + passive
    """
    if cash_rate <= 0:
        return 0.5

    # Primary signal: cash rate relative to the "pro optimal" of ~18%
    # Players well above 18% are being passive (or in very soft fields)
    # Players near/below 18% with positive results are aggressive
    #
    # Map: 10% CR -> 0.75 agg, 18% -> 0.55, 25% -> 0.40, 35% -> 0.25, 45% -> 0.15
    # This reflects that higher CR = more passive in tournament poker
    agg_from_cr = max(0.10, min(0.85, 0.80 - (cash_rate - 0.10) * 1.8))

    # Secondary signal: FT rate per cash (only if field-normalized)
    # In normalized data, high FT/cash = aggressive (big stacks going deep)
    # In raw data from small fields, this is unreliable so we weight it less
    ft_to_cash = ft_rate / cash_rate if cash_rate > 0 else 0.3
    agg_from_ft = max(0.10, min(0.90, (ft_to_cash - 0.15) / 0.50))

    # Tertiary signal: top-3 conversion at FT
    # Aggressive players convert more FTs to wins (big stack advantage)
    agg_from_t3 = max(0.10, min(0.90, (top3_conv - 0.20) / 0.50))

    if field_normalized:
        # FT data is reliable — use all signals
        raw = agg_from_cr * 0.45 + agg_from_ft * 0.35 + agg_from_t3 * 0.20
    else:
        # FT data may be inflated by small fields — lean heavily on CR
        raw = agg_from_cr * 0.70 + agg_from_ft * 0.15 + agg_from_t3 * 0.15

    return max(0.05, min(0.95, raw))


# ============================================================
# CASH RATE DEGRADATION MODEL
# ============================================================

def degradation_per_step(current_cr):
    """Cash rate drop for ONE tier jump up."""
    if current_cr >= 0.38:
        return 0.020 + (0.45 - current_cr) * 0.03
    elif current_cr >= 0.30:
        return 0.030 + (0.38 - current_cr) * 0.06
    elif current_cr >= 0.25:
        return 0.040 + (0.30 - current_cr) * 0.08
    elif current_cr >= 0.20:
        return 0.050 + (0.25 - current_cr) * 0.12
    elif current_cr >= 0.15:
        return 0.055 + (0.20 - current_cr) * 0.15
    else:
        return 0.065 + (0.15 - current_cr) * 0.20


def project_cash_rate(current_cr, steps_up, aggression=0.5):
    """Project cash rate after N steps up. Each step compounds."""
    cr = current_cr
    for _ in range(steps_up):
        drop = degradation_per_step(cr)
        cr = max(0.03, cr - drop)
    return cr


def project_ft_rate(current_ft_rate, current_cr, new_cr):
    """FT rate degrades proportionally to cash rate, with slight extra penalty."""
    if current_cr == 0:
        return 0
    ratio = new_cr / current_cr
    return max(0.005, current_ft_rate * ratio * 0.92)


# ============================================================
# FIELD SIZE NORMALIZATION
# ============================================================
# Problem: a "4th place" in a 100-person nightly is ~top 4% (min-cash zone)
# but "4th place" in a 250-person festival is ~top 1.6% (deep FT run).
# Hendon Mob doesn't distinguish, so players in small fields have
# inflated FT counts.
#
# Fix: convert each placement to a percentile in the original field,
# then re-map to what that percentile means in a target field size.

def normalize_placements(player, avg_field_size, target_field_size=250):
    """Normalize a player's placement distribution from their actual
    field sizes to a target field size.

    In a field of N with 12.5% paid:
      - Places paid = N * 0.125
      - FT = top 9 (fixed)
      - Min-cash zone = places 10 through places_paid

    A "5th place" in a 100-person field (top 5%) maps to ~13th in a
    250-person field. That 13th is a min-cash, not a FT finish.
    """
    t = player["tournaments"]
    paid_orig = max(9, round(avg_field_size * 0.125))
    paid_target = max(9, round(target_field_size * 0.125))

    # Original placement counts
    orig_1st = player["1st"]
    orig_2nd = player["2nd"]
    orig_3rd = player["3rd"]
    orig_4_9 = player["4th-9th"]
    orig_10_plus = player["10th+"]
    total_cashes = orig_1st + orig_2nd + orig_3rd + orig_4_9 + orig_10_plus

    if total_cashes == 0 or avg_field_size == target_field_size:
        return player.copy()

    # For each placement bucket, compute the percentile in the original field
    # then decide where it lands in the target field.
    #
    # 1st-3rd: These are absolute skill — top 3 is top 3 regardless of field.
    # BUT in a smaller field, getting to top 3 is easier (fewer opponents).
    # Scale by ratio of fields: a 1st in 100 players is ~2.5x easier than
    # 1st in 250 players. We apply a conversion rate.
    field_ratio = target_field_size / avg_field_size  # e.g., 250/100 = 2.5

    # Top-3 finishes: harder to achieve in bigger fields
    # A player who wins 5% of 100-person fields would win ~2.5% of 250-person fields
    # But it's not purely linear — skill compounds in deeper fields
    # Use sqrt scaling: harder but not proportionally harder
    top3_scale = 1.0 / (field_ratio ** 0.35)  # e.g., 2.5^0.35 = ~1.38, so 1/1.38 = 0.72

    new_1st = max(0, round(orig_1st * top3_scale))
    new_2nd = max(0, round(orig_2nd * top3_scale))
    new_3rd = max(0, round(orig_3rd * top3_scale))

    # 4th-9th: This is where the big correction happens.
    # In a 100-person field, 4th-9th = top 4-9% = basically bubble/min-cash
    # In a 250-person field, 4th-9th = top 1.6-3.6% = legit deep run
    #
    # What fraction of the original 4th-9th finishes would actually be
    # 4th-9th in the larger field vs falling to 10th+?
    #
    # Percentile of "9th place" in orig field:
    pct_9th_orig = 9.0 / avg_field_size       # e.g., 9/100 = 9%
    pct_9th_target = 9.0 / target_field_size   # e.g., 9/250 = 3.6%

    # What fraction of the original 4th-9th range falls within the target's
    # 4th-9th percentile? The original 4th-9th spans percentile
    # (3/field) to (9/field). The target's 4th-9th spans (3/250) to (9/250).
    # Overlap determines how many stay as FT finishes.
    if pct_9th_orig > 0:
        # Fraction of original 4-9th that would still be 4-9th in bigger field
        ft_retention = min(1.0, pct_9th_target / pct_9th_orig)
    else:
        ft_retention = 1.0

    new_4_9 = max(0, round(orig_4_9 * ft_retention * top3_scale))

    # Everything else becomes 10th+
    new_top_finishes = new_1st + new_2nd + new_3rd + new_4_9
    # Total cashes stays the same (cash rate doesn't change with field size,
    # it's a player skill metric). But more cashes land in 10th+.
    new_10_plus = max(0, total_cashes - new_top_finishes)

    # Build normalized player
    normalized = player.copy()
    normalized["1st"] = new_1st
    normalized["2nd"] = new_2nd
    normalized["3rd"] = new_3rd
    normalized["4th-9th"] = new_4_9
    normalized["10th+"] = new_10_plus

    # Recalculate rates
    new_ft = new_1st + new_2nd + new_3rd + new_4_9
    normalized["ft_rate"] = new_ft / t if t > 0 else 0
    normalized["top3_conv"] = (new_1st + new_2nd + new_3rd) / max(1, new_ft)
    # Cash rate stays the same — how often you cash doesn't change

    normalized["_raw"] = {
        "1st": orig_1st, "2nd": orig_2nd, "3rd": orig_3rd,
        "4th-9th": orig_4_9, "10th+": orig_10_plus,
    }
    normalized["_field_norm"] = f"{avg_field_size} -> {target_field_size}"

    return normalized


# Field size estimates for our players' typical games
# Wynn $250 (150+100): ~100 players
# $300: ~250 players
# Fri/Sat/Sun: 250+ regardless
# Weighted avg depends on how much each player plays weekday vs weekend

FIELD_ESTIMATES = {
    "Nicholas": 150,   # mix of weekday/weekend
    "Vincent": 120,    # nightly grinder, mostly weekday
    "Bryan": 130,      # mix
    "Nikko": 130,      # mix
    "Becker": 150,     # volume player, wide mix
    "Dan": 100,        # mostly small weekday nightlies
    "Frankie": 120,    # mostly small fields
}

# Target field size = typical $600-$1,500 tournament
TARGET_FIELD = 250


# ============================================================
# REAL PLAYER DATA
# ============================================================
REAL_PLAYERS = [
    {
        "name": "Bryan",
        "tournaments": 812,
        "cash_rate": 0.250,
        "1st": 23, "2nd": 20, "3rd": 15, "4th-9th": 50, "10th+": 95,
        "ft_rate": 0.133,
        "top3_conv": 0.537,
        "current_tier": "low",
        "aggression": 0.72,  # most aggressive — trades CR for big stacks
    },
    {
        "name": "Becker",
        "tournaments": 3961,
        "cash_rate": 0.180,
        "1st": 46, "2nd": 41, "3rd": 43, "4th-9th": 162, "10th+": 421,
        "ft_rate": 0.0737,
        "top3_conv": 0.445,
        "current_tier": "low",
        "aggression": 0.68,  # close behind Bryan
    },
    {
        "name": "Nicholas",
        "tournaments": 110,
        "cash_rate": 0.345,
        "1st": 5, "2nd": 4, "3rd": 3, "4th-9th": 12, "10th+": 14,
        "ft_rate": 0.218,
        "top3_conv": 0.50,
        "current_tier": "low",
        "aggression": 0.58,  # slightly above balanced
    },
    {
        "name": "Vincent",
        "tournaments": 630,
        "cash_rate": 0.338,
        "1st": 25, "2nd": 20, "3rd": 18, "4th-9th": 55, "10th+": 95,
        "ft_rate": 0.187,
        "top3_conv": 0.534,
        "current_tier": "low",
        "aggression": 0.55,  # very close to Nicholas
    },
    {
        "name": "Nikko",
        "tournaments": 491,
        "cash_rate": 0.320,
        "1st": 12, "2nd": 9, "3rd": 14, "4th-9th": 32, "10th+": 90,
        "ft_rate": 0.1365,
        "top3_conv": 0.522,
        "current_tier": "low",
        "aggression": 0.40,  # noticeably behind Vince/Nicholas but well ahead of bottom
    },
    {
        "name": "Frankie",
        "tournaments": 239,
        "cash_rate": 0.209,
        "1st": 1, "2nd": 2, "3rd": 1, "4th-9th": 3, "10th+": 43,
        "ft_rate": 0.0293,
        "top3_conv": 0.571,
        "current_tier": "low",
        "aggression": 0.22,  # very passive — min-cash machine
    },
    {
        "name": "Dan",
        "tournaments": 322,
        "cash_rate": 0.137,  # 44 cashes / 322 tournaments
        "1st": 2, "2nd": 3, "3rd": 4, "4th-9th": 20, "10th+": 15,
        "ft_rate": 0.090,  # 29 FTs (2+3+4+20) / 322
        "top3_conv": 0.310,  # 9 top-3 / 29 FT
        "current_tier": "low",
        "aggression": 0.18,  # most passive in the group
    },
]


# ============================================================
# CORE CALCULATIONS
# ============================================================

def calc_ev(placements, tournaments, payouts, buyin):
    """EV per entry given placement counts and payout structure."""
    gross = (
        placements["1st"] / tournaments * payouts["1st"]
        + placements["2nd"] / tournaments * payouts["2nd"]
        + placements["3rd"] / tournaments * payouts["3rd"]
        + placements["4th-9th"] / tournaments * payouts["4th-9th"]
        + placements["10th+"] / tournaments * payouts["10th+"]
    )
    return gross - buyin


def project_placements(player, new_cr, new_ft_rate, aggression=None):
    """Project placement distribution at a new stake level.

    Aggression shifts the distribution:
    - High aggression: fewer min-cashes, more 1st places
    - Low aggression: more min-cashes, fewer deep runs
    """
    t = player["tournaments"]
    agg = aggression if aggression is not None else player.get("aggression", 0.5)

    # Apply aggression to the projected rates
    adj_cr, adj_ft, adj_t3, stack_mult = apply_aggression(new_cr, new_ft_rate,
                                                           player.get("top3_conv", 0.33), agg)

    new_total_cashes = max(1, round(t * adj_cr))
    new_total_ft = max(0, round(t * adj_ft))
    new_total_ft = min(new_total_ft, new_total_cashes)

    new_top3 = max(0, round(new_total_ft * adj_t3))
    new_top3 = min(new_top3, new_total_ft)

    # Top-3 distribution shifts with aggression
    # Aggressive players win MORE 1sts relative to 2nd/3rd
    # because bigger stacks dominate heads-up
    top3_sum = player["1st"] + player["2nd"] + player["3rd"]
    if top3_sum > 0:
        base_first_pct = player["1st"] / top3_sum
        base_second_pct = player["2nd"] / top3_sum
    else:
        base_first_pct, base_second_pct = 0.30, 0.33

    # Aggression skews toward 1st: big stack = more wins
    delta = agg - 0.5
    first_pct = min(0.60, base_first_pct + delta * 0.15)
    third_pct = max(0.15, (1 - base_first_pct - base_second_pct) - delta * 0.10)
    second_pct = max(0.15, 1.0 - first_pct - third_pct)

    new_1st = max(0, round(new_top3 * first_pct))
    new_2nd = max(0, round(new_top3 * second_pct))
    new_3rd = max(0, new_top3 - new_1st - new_2nd)
    new_4th_9th = max(0, new_total_ft - new_1st - new_2nd - new_3rd)
    new_10th_plus = max(0, new_total_cashes - new_1st - new_2nd - new_3rd - new_4th_9th)

    return {
        "1st": new_1st, "2nd": new_2nd, "3rd": new_3rd,
        "4th-9th": new_4th_9th, "10th+": new_10th_plus,
    }


def evaluate_player_full(player):
    """Evaluate a player across all 4 tiers."""
    t = player["tournaments"]
    name = player["name"]
    current_tier = player["current_tier"]
    current_idx = TIER_ORDER.index(current_tier)
    hrs_per_entry = 6.5

    current_placements = {
        "1st": player["1st"], "2nd": player["2nd"], "3rd": player["3rd"],
        "4th-9th": player["4th-9th"], "10th+": player["10th+"],
    }

    agg = player.get("aggression", 0.5)
    _, _, _, stack_mult = apply_aggression(player["cash_rate"], player["ft_rate"],
                                           player.get("top3_conv", 0.33), agg)

    results = {"Name": name, "Tournaments": t, "Current Tier": TIERS[current_tier]["label"],
               "Aggression": round(agg, 2), "Stack@MinCash": f"{stack_mult:.0f}x"}

    tier_evs = {}
    tier_hrs = {}

    for i, tier_key in enumerate(TIER_ORDER):
        tier = TIERS[tier_key]
        label = tier["label"]

        if tier_key == current_tier:
            cr = player["cash_rate"]
            ft = player["ft_rate"]
            placements = current_placements
        else:
            steps = i - current_idx
            if steps > 0:
                cr = project_cash_rate(player["cash_rate"], steps, aggression=agg)
                ft = project_ft_rate(player["ft_rate"], player["cash_rate"], cr)
                placements = project_placements(player, cr, ft, aggression=agg)
            else:
                cr = min(0.55, player["cash_rate"] + abs(steps) * 0.04)
                ft = min(0.35, player["ft_rate"] * (cr / player["cash_rate"]) * 1.05)
                placements = project_placements(player, cr, ft, aggression=agg)

        ev = calc_ev(placements, t, tier["payouts"], tier["buyin"])
        roi = (ev / tier["buyin"]) * 100
        hourly = ev / hrs_per_entry

        tier_evs[tier_key] = ev
        tier_hrs[tier_key] = hourly

        results[f"CR ({label})"] = round(cr, 4)
        results[f"FT ({label})"] = round(ft, 4)
        results[f"EV ({label})"] = round(ev, 0)
        results[f"ROI ({label})"] = round(roi, 1)
        results[f"$/hr ({label})"] = round(hourly, 0)

    # Find the tier with the highest hourly
    best_tier = max(TIER_ORDER, key=lambda k: tier_hrs[k])
    best_label = TIERS[best_tier]["label"]
    best_hr = tier_hrs[best_tier]

    # Recommendation logic
    current_hr = tier_hrs[current_tier]
    current_ev = tier_evs[current_tier]

    if current_ev <= 0:
        # Check if profitable at a higher tier (payout structure can overcome thin edge)
        profitable_higher = [(k, tier_evs[k], tier_hrs[k]) for k in TIER_ORDER
                             if TIER_ORDER.index(k) > current_idx and tier_evs[k] > 0]
        if profitable_higher:
            best_higher = max(profitable_higher, key=lambda x: x[2])
            rec = f"MOVE UP to {TIERS[best_higher[0]]['label']}"
            reason = (f"Breakeven/losing at $300 (${current_ev:+,.0f}) but payout structure at "
                     f"{TIERS[best_higher[0]]['label']} is profitable (${best_higher[1]:+,.0f}/entry, "
                     f"${best_higher[2]:,.0f}/hr).")
        else:
            rec = "MOVE DOWN"
            reason = f"Losing at current level (${current_ev:+,.0f}/entry) and no higher tier is profitable."
    elif best_tier == current_tier:
        rec = f"STAY at {TIERS[current_tier]['label']}"
        reason = f"Best hourly (${current_hr:,.0f}/hr) is already at current level."
    else:
        best_idx = TIER_ORDER.index(best_tier)
        # Check each step up to see if it's profitable
        step_up_viable = True
        for step_key in TIER_ORDER[current_idx + 1:best_idx + 1]:
            if tier_evs[step_key] <= 0:
                step_up_viable = False
                # Find highest viable tier
                viable_tiers = [k for k in TIER_ORDER if tier_evs[k] > 0]
                if viable_tiers:
                    best_viable = max(viable_tiers, key=lambda k: tier_hrs[k])
                    if best_viable != current_tier:
                        rec = f"MOVE UP to {TIERS[best_viable]['label']}"
                        reason = f"Profitable up to {TIERS[best_viable]['label']} (${tier_hrs[best_viable]:,.0f}/hr) but loses money at {TIERS[step_key]['label']}."
                    else:
                        rec = f"STAY at {TIERS[current_tier]['label']}"
                        reason = f"Best hourly at current level. Goes negative at {TIERS[step_key]['label']}."
                else:
                    rec = "MOVE DOWN"
                    reason = "Negative EV at all levels."
                break

        if step_up_viable:
            hr_gain = best_hr / max(1, current_hr)
            if hr_gain > 1.3:
                rec = f"MOVE UP to {best_label}"
                reason = f"Hourly jumps {hr_gain - 1:.0%} (${current_hr:,.0f} -> ${best_hr:,.0f}/hr). Profitable at all steps."
            elif hr_gain > 1.0:
                rec = f"CONSIDER {best_label}"
                reason = f"Modest hourly gain (${current_hr:,.0f} -> ${best_hr:,.0f}/hr). Need ${100 * TIERS[best_tier]['buyin']:,} bankroll."
            else:
                rec = f"STAY at {TIERS[current_tier]['label']}"
                reason = f"Best hourly at current level despite being profitable higher."

    results["Best Tier"] = best_label
    results["Best $/hr"] = round(best_hr, 0)
    results["Recommendation"] = rec
    results["Reasoning"] = reason

    return results


# ============================================================
# EXCEL OUTPUT
# ============================================================

def create_workbook(player_results):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stake Move Analysis"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    rec_colors = {}
    for r in player_results:
        rec = r.get("Recommendation", "")
        if "MOVE UP" in rec:
            rec_colors[rec] = "2d6a4f"
        elif "CONSIDER" in rec:
            rec_colors[rec] = "52b788"
        elif "STAY" in rec:
            rec_colors[rec] = "e9c46a"
        elif "MOVE DOWN" in rec:
            rec_colors[rec] = "e76f51"

    headers = list(player_results[0].keys())

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, player in enumerate(player_results, 2):
        rec = player.get("Recommendation", "")
        for col_idx, h in enumerate(headers, 1):
            val = player[h]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", wrap_text=(h == "Reasoning"))

            if h == "Recommendation":
                color = rec_colors.get(rec, "e9c46a")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            if "CR " in h or "FT " in h:
                cell.number_format = "0.0%"
            elif "EV " in h or "$/hr" in h or "Best $/hr" in h:
                cell.number_format = '$#,##0'
            elif "ROI" in h:
                cell.number_format = '0.0"%"'

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    for col, h in enumerate(headers, 1):
        if h == "Reasoning":
            ws.column_dimensions[get_column_letter(col)].width = 65
        elif h in ("Name", "Recommendation", "Best Tier", "Current Tier"):
            ws.column_dimensions[get_column_letter(col)].width = 22
    ws.freeze_panes = "B2"

    # --- Sheet 2: Payout Comparison ---
    ws2 = wb.create_sheet("Payout Structures")

    tier_keys = TIER_ORDER
    payout_headers = ["Place"] + [TIERS[k]["label"] for k in tier_keys]
    for i in range(len(tier_keys) - 1):
        payout_headers.append(f"{TIERS[tier_keys[i]]['label']} -> {TIERS[tier_keys[i+1]]['label']}")

    for col, h in enumerate(payout_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    places = ["1st", "2nd", "3rd", "4th-9th", "10th+"]
    place_labels = ["1st Place", "2nd Place", "3rd Place", "4th-9th (avg)", "10th+ (min cash avg)"]

    for row_idx, (place, label) in enumerate(zip(places, place_labels), 2):
        vals = [label]
        payouts_by_tier = []
        for k in tier_keys:
            p = TIERS[k]["payouts"][place]
            payouts_by_tier.append(p)
            vals.append(p)
        for i in range(len(payouts_by_tier) - 1):
            vals.append(round(payouts_by_tier[i + 1] / max(1, payouts_by_tier[i]), 1))

        for col_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if 1 < col_idx <= len(tier_keys) + 1:
                cell.number_format = '$#,##0'
            elif col_idx > len(tier_keys) + 1:
                cell.number_format = '0.0"x"'

    # Buy-in row
    row_idx = len(places) + 2
    vals = ["Buy-in (avg)"] + [TIERS[k]["buyin"] for k in tier_keys]
    for i in range(len(tier_keys) - 1):
        vals.append(round(TIERS[tier_keys[i + 1]]["buyin"] / max(1, TIERS[tier_keys[i]]["buyin"]), 1))
    for col_idx, val in enumerate(vals, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        if 1 < col_idx <= len(tier_keys) + 1:
            cell.number_format = '$#,##0'

    # Bankroll row
    row_idx += 1
    vals = ["Bankroll (100 BI)"] + [TIERS[k]["buyin"] * 100 for k in tier_keys] + [""] * (len(tier_keys) - 1)
    for col_idx, val in enumerate(vals, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
        if 1 < col_idx <= len(tier_keys) + 1 and val:
            cell.number_format = '$#,##0'

    # Sample size row
    row_idx += 1
    vals = ["Sample size"] + [TIERS[k]["sample_size"] for k in tier_keys] + [""] * (len(tier_keys) - 1)
    for col_idx, val in enumerate(vals, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(payout_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 20

    # --- Sheet 3: Degradation Table ---
    ws3 = wb.create_sheet("Cash Rate Degradation")

    ws3.cell(row=1, column=1, value="Cash Rate Projections Across All Tiers").font = Font(bold=True, size=13)

    deg_headers = ["Current CR at $300"] + [f"Projected at {TIERS[k]['label']}" for k in TIER_ORDER[1:]] + \
                  [f"Drop to {TIERS[k]['label']}" for k in TIER_ORDER[1:]]
    for col, h in enumerate(deg_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    test_rates = [0.10, 0.12, 0.15, 0.18, 0.20, 0.22, 0.25, 0.28, 0.30, 0.33, 0.35, 0.38, 0.40, 0.45]
    for row_idx, cr in enumerate(test_rates, 4):
        vals = [cr]
        for steps in range(1, len(TIER_ORDER)):
            vals.append(project_cash_rate(cr, steps))
        for steps in range(1, len(TIER_ORDER)):
            vals.append(cr - project_cash_rate(cr, steps))

        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=round(val, 4))
            cell.number_format = "0.0%"
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(deg_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    return wb


def main():
    print("=== Stake Move Model (4 Tiers — Real Data) ===\n")

    print("Payout structures:")
    for key in TIER_ORDER:
        tier = TIERS[key]
        p = tier["payouts"]
        print(f"  {tier['label']:>15s} (avg ${tier['buyin']:,}, n={tier['sample_size']}): "
              f"1st=${p['1st']:,}  2nd=${p['2nd']:,}  3rd=${p['3rd']:,}  "
              f"4-9th=${p['4th-9th']:,}  10+=${p['10th+']:,}")

    # Normalize players for field size
    print("\n--- Field Size Normalization ---\n")
    print(f"{'Name':>12s} | {'Field':>5s} | {'Raw 1/2/3/4-9/10+':>25s} | {'Norm 1/2/3/4-9/10+':>25s} | {'FT%':>5s} | {'T3@FT':>5s}")
    print("-" * 100)
    normalized_players = []
    for p in REAL_PLAYERS:
        field = FIELD_ESTIMATES.get(p["name"], 120)
        np_ = normalize_placements(p, avg_field_size=field, target_field_size=TARGET_FIELD)
        normalized_players.append(np_)
        raw = f"{p['1st']}/{p['2nd']}/{p['3rd']}/{p['4th-9th']}/{p['10th+']}"
        norm = f"{np_['1st']}/{np_['2nd']}/{np_['3rd']}/{np_['4th-9th']}/{np_['10th+']}"
        print(f"{p['name']:>12s} | {field:>5d} | {raw:>25s} | {norm:>25s} | {np_['ft_rate']:>4.1%} | {np_['top3_conv']:>4.1%}")

    print("\n--- Player Evaluations ---\n")
    header = f"{'Name':>12s} {'Agg':>4s} {'Stk':>4s}"
    for k in TIER_ORDER:
        label = TIERS[k]['label']
        header += f" | {'CR':>6s} {'EV':>8s} {'$/hr':>6s}"
    header += " | Recommendation"
    print(header)
    print("-" * len(header))

    results = []
    for p in normalized_players:
        r = evaluate_player_full(p)
        results.append(r)

        line = f"{r['Name']:>12s} {r['Aggression']:>4.2f} {r['Stack@MinCash']:>4s}"
        for k in TIER_ORDER:
            label = TIERS[k]['label']
            line += f" | {r[f'CR ({label})']:>5.1%} ${r[f'EV ({label})']:>+7,.0f} ${r[f'$/hr ({label})']:>5,.0f}"
        line += f" | {r['Recommendation']}"
        print(line)

    output_path = "c:/Users/nicho/OneDrive/Desktop/code/Poker_videos/stake_move_analysis.xlsx"
    wb = create_workbook(results)
    wb.save(output_path)
    print(f"\nSaved -> {output_path}")


if __name__ == "__main__":
    main()
