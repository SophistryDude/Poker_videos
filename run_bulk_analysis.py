"""
Run all 100K synthetic players through the 4-tier stake move model.
"""

import csv
import time
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

from stake_move_model import (
    TIERS, TIER_ORDER, project_cash_rate, project_ft_rate,
    project_placements, calc_ev, apply_aggression,
)

HRS_PER_ENTRY = 6.5

PLAYER_TIERS = [
    "Brand New", "Bad Reg", "Average Player", "Slightly Profitable",
    "Good Reg", "Great Reg", "Low Level Pro", "Mid Level Pro",
    "High Level Pro", "Best in the World",
]


def evaluate_synthetic(p):
    """Evaluate a synthetic player across all 4 payout tiers."""
    t = p["tournaments"]
    if t == 0:
        return None

    agg = p.get("aggression", 0.5)

    current_placements = {
        "1st": p["1st"], "2nd": p["2nd"], "3rd": p["3rd"],
        "4th-9th": p["4th-9th"], "10th+": p["10th+"],
    }

    tier_evs = {}
    tier_hrs = {}
    tier_crs = {}

    for i, tier_key in enumerate(TIER_ORDER):
        tier = TIERS[tier_key]
        if i == 0:
            # Current level — use actual stats
            cr = p["cash_rate"]
            ft = p["ft_rate"]
            placements = current_placements
        else:
            cr = project_cash_rate(p["cash_rate"], i, aggression=agg)
            ft = project_ft_rate(p["ft_rate"], p["cash_rate"], cr)
            placements = project_placements(p, cr, ft, aggression=agg)

        ev = calc_ev(placements, t, tier["payouts"], tier["buyin"])
        tier_evs[tier_key] = ev
        tier_hrs[tier_key] = ev / HRS_PER_ENTRY
        tier_crs[tier_key] = cr

    # Recommendation logic (mirrors evaluate_player_full)
    current_ev = tier_evs["low"]
    current_hr = tier_hrs["low"]

    if current_ev <= 0:
        profitable_higher = [(k, tier_evs[k], tier_hrs[k]) for k in TIER_ORDER[1:]
                             if tier_evs[k] > 0]
        if profitable_higher:
            best = max(profitable_higher, key=lambda x: x[2])
            rec = f"MOVE UP to {TIERS[best[0]]['label']}"
        else:
            rec = "MOVE DOWN"
    else:
        best_tier = max(TIER_ORDER, key=lambda k: tier_hrs[k])
        best_hr = tier_hrs[best_tier]

        if best_tier == "low":
            rec = "STAY at $0-$300"
        else:
            # Check all steps up are profitable
            best_idx = TIER_ORDER.index(best_tier)
            step_up_viable = True
            ceiling = best_tier
            for step_key in TIER_ORDER[1:best_idx + 1]:
                if tier_evs[step_key] <= 0:
                    step_up_viable = False
                    # Find highest viable
                    viable = [k for k in TIER_ORDER if tier_evs[k] > 0]
                    if viable:
                        best_viable = max(viable, key=lambda k: tier_hrs[k])
                        if best_viable != "low":
                            rec = f"MOVE UP to {TIERS[best_viable]['label']}"
                        else:
                            rec = "STAY at $0-$300"
                    else:
                        rec = "MOVE DOWN"
                    break

            if step_up_viable:
                hr_gain = best_hr / max(1, current_hr)
                if hr_gain > 1.3:
                    rec = f"MOVE UP to {TIERS[best_tier]['label']}"
                elif hr_gain > 1.0:
                    rec = f"CONSIDER {TIERS[best_tier]['label']}"
                else:
                    rec = "STAY at $0-$300"

    return {
        "id": p["id"],
        "tier": p["tier"],
        "tournaments": t,
        "cash_rate": p["cash_rate"],
        "ft_rate": p["ft_rate"],
        "top3_conv": p["top3_conv"],
        "ev_low": round(tier_evs["low"], 0),
        "hr_low": round(tier_hrs["low"], 0),
        "cr_mid_low": round(tier_crs["mid_low"], 4),
        "ev_mid_low": round(tier_evs["mid_low"], 0),
        "hr_mid_low": round(tier_hrs["mid_low"], 0),
        "cr_mid": round(tier_crs["mid"], 4),
        "ev_mid": round(tier_evs["mid"], 0),
        "hr_mid": round(tier_hrs["mid"], 0),
        "cr_high": round(tier_crs["high"], 4),
        "ev_high": round(tier_evs["high"], 0),
        "hr_high": round(tier_hrs["high"], 0),
        "rec": rec,
    }


def load_synthetic_players(csv_path):
    players = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            players.append({
                "id": int(row["id"]),
                "tier": row["tier"],
                "tournaments": int(row["tournaments"]),
                "1st": int(row["1st"]),
                "2nd": int(row["2nd"]),
                "3rd": int(row["3rd"]),
                "4th-9th": int(row["4th-9th"]),
                "10th+": int(row["10th+"]),
                "cash_rate": float(row["cash_rate"]),
                "ft_rate": float(row["ft_rate"]),
                "top3_conv": float(row["top3_conv"]),
                "aggression": float(row.get("aggression", 0.5)),
            })
    return players


def build_workbook(results):
    wb = openpyxl.Workbook()

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    halign = Alignment(horizontal="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    tier_colors = {
        "Brand New": "FFE0E0", "Bad Reg": "FFD0C0", "Average Player": "FFE8B0",
        "Slightly Profitable": "FFFFC0", "Good Reg": "D8F0D8", "Great Reg": "B0E0B0",
        "Low Level Pro": "B0D8F0", "Mid Level Pro": "90C0E8",
        "High Level Pro": "C0B0F0", "Best in the World": "FFD700",
    }

    # Gather all unique recs
    all_recs = sorted(set(r["rec"] for r in results),
                      key=lambda x: (0 if "DOWN" in x else 1 if "STAY" in x else 2 if "CONSIDER" in x else 3))

    def rec_color(rec):
        if "MOVE DOWN" in rec: return "e76f51"
        if "STAY" in rec: return "e9c46a"
        if "CONSIDER" in rec: return "52b788"
        if "$1,500" in rec: return "1b4332"
        if "$601" in rec: return "2d6a4f"
        if "$301" in rec: return "40916c"
        return "888888"

    # ── Sheet 1: Recommendations by Player Tier ──
    ws1 = wb.active
    ws1.title = "Recommendations by Tier"

    s1_headers = ["Player Tier", "Count"] + all_recs + ["% Should Move Up"]
    for col, h in enumerate(s1_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    for row_idx, tn in enumerate(PLAYER_TIERS, 2):
        tier_results = [r for r in results if r["tier"] == tn]
        counts = Counter(r["rec"] for r in tier_results)
        total = len(tier_results)
        move_up = sum(v for k, v in counts.items() if "MOVE UP" in k)
        mu_pct = move_up / max(1, total)

        fill = PatternFill(start_color=tier_colors.get(tn, "FFFFFF"),
                          end_color=tier_colors.get(tn, "FFFFFF"), fill_type="solid")

        vals = [tn, total] + [counts.get(r, 0) for r in all_recs] + [round(mu_pct, 4)]
        for col_idx, val in enumerate(vals, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if col_idx == len(vals):
                cell.number_format = "0.0%"
            elif col_idx > 2 and col_idx < len(vals):
                rc_idx = col_idx - 3
                if rc_idx < len(all_recs) and val > 0:
                    rc = rec_color(all_recs[rc_idx])
                    cell.fill = PatternFill(start_color=rc, end_color=rc, fill_type="solid")
                    cell.font = Font(color="FFFFFF")

    # Total row
    row_idx = len(PLAYER_TIERS) + 2
    total_counts = Counter(r["rec"] for r in results)
    total_mu = sum(v for k, v in total_counts.items() if "MOVE UP" in k) / max(1, len(results))
    total_vals = ["TOTAL", len(results)] + [total_counts.get(r, 0) for r in all_recs] + [round(total_mu, 4)]
    for col_idx, val in enumerate(total_vals, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        if col_idx == len(total_vals):
            cell.number_format = "0.0%"

    for col in range(1, len(s1_headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 18
    ws1.column_dimensions["A"].width = 22
    ws1.freeze_panes = "B2"

    # ── Sheet 2: Decision Boundaries ──
    ws2 = wb.create_sheet("Decision Boundaries")
    ws2.cell(row=1, column=1, value="Cash Rate Thresholds by Recommendation").font = Font(bold=True, size=13)

    bound_headers = ["Recommendation", "Count", "Min CR", "P10 CR", "Median CR", "P90 CR", "Max CR",
                     "Avg EV@$300", "Avg EV@$600", "Avg EV@$1K", "Avg EV@$1.8K"]
    for col, h in enumerate(bound_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    for row_idx, rec in enumerate(all_recs, 4):
        recs = [r for r in results if r["rec"] == rec]
        if not recs:
            continue
        crs = [r["cash_rate"] for r in recs]
        vals = [
            rec, len(recs),
            round(min(crs), 4), round(np.percentile(crs, 10), 4),
            round(np.median(crs), 4), round(np.percentile(crs, 90), 4),
            round(max(crs), 4),
            round(np.mean([r["ev_low"] for r in recs]), 0),
            round(np.mean([r["ev_mid_low"] for r in recs]), 0),
            round(np.mean([r["ev_mid"] for r in recs]), 0),
            round(np.mean([r["ev_high"] for r in recs]), 0),
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if col_idx == 1:
                cell.fill = PatternFill(start_color=rec_color(rec), end_color=rec_color(rec), fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            elif col_idx in (3, 4, 5, 6, 7):
                cell.number_format = "0.0%"
            elif col_idx >= 8:
                cell.number_format = '$#,##0'

    for col in range(1, len(bound_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18
    ws2.column_dimensions["A"].width = 28

    # ── Sheet 3: Where Each Tier's Best Level Is ──
    ws3 = wb.create_sheet("Optimal Level by Tier")
    ws3.cell(row=1, column=1, value="Where is the highest hourly for each player tier?").font = Font(bold=True, size=13)

    payout_labels = [TIERS[k]["label"] for k in TIER_ORDER]
    opt_headers = ["Player Tier", "Count"] + [f"Best = {l}" for l in payout_labels]
    for col, h in enumerate(opt_headers, 1):
        cell = ws3.cell(row=3, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    for row_idx, tn in enumerate(PLAYER_TIERS, 4):
        tr = [r for r in results if r["tier"] == tn]
        fill = PatternFill(start_color=tier_colors.get(tn, "FFFFFF"),
                          end_color=tier_colors.get(tn, "FFFFFF"), fill_type="solid")

        # Find which payout tier gives best hourly for each player
        best_counts = Counter()
        for r in tr:
            hrs = {"low": r["hr_low"], "mid_low": r["hr_mid_low"],
                   "mid": r["hr_mid"], "high": r["hr_high"]}
            best = max(hrs, key=lambda k: hrs[k])
            best_counts[best] += 1

        vals = [tn, len(tr)] + [best_counts.get(k, 0) for k in TIER_ORDER]
        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(opt_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 18
    ws3.column_dimensions["A"].width = 22

    # ── Sheet 4: Edge Cases ──
    ws4 = wb.create_sheet("Edge Cases (200)")

    s4_headers = ["Type", "Player Tier", "Tourns", "Cash Rate", "FT Rate", "T3 Conv",
                  "EV@$300", "EV@$600", "EV@$1K", "EV@$1.8K", "Recommendation"]
    for col, h in enumerate(s4_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    edge_cases = []

    # High CR but best level is $300
    stay_high = [r for r in results if "STAY" in r["rec"] and r["cash_rate"] >= 0.25]
    stay_high.sort(key=lambda r: -r["cash_rate"])
    for r in stay_high[:30]:
        edge_cases.append(("High CR / STAY", r))

    # Low CR but MOVE UP to $1,500+
    moveup_top_low_cr = [r for r in results if "$1,500" in r["rec"] and r["cash_rate"] <= 0.25]
    moveup_top_low_cr.sort(key=lambda r: r["cash_rate"])
    for r in moveup_top_low_cr[:30]:
        edge_cases.append(("Low CR / UP to $1.5K", r))

    # Losing at $300 but profitable higher
    losing_low_profit_high = [r for r in results if r["ev_low"] <= 0 and r["ev_mid_low"] > 0]
    losing_low_profit_high.sort(key=lambda r: r["ev_mid_low"])
    for r in losing_low_profit_high[-30:]:
        edge_cases.append(("Losing low / UP higher", r))

    # Small sample (<80 tourns) with MOVE UP
    small_up = [r for r in results if "MOVE UP" in r["rec"] and r["tournaments"] < 80]
    small_up.sort(key=lambda r: r["tournaments"])
    for r in small_up[:30]:
        edge_cases.append(("Small Sample Hero", r))

    # Profitable at all 4 tiers
    all_profit = [r for r in results if r["ev_low"] > 0 and r["ev_mid_low"] > 0
                  and r["ev_mid"] > 0 and r["ev_high"] > 0]
    all_profit.sort(key=lambda r: -r["ev_high"])
    for r in all_profit[:30]:
        edge_cases.append(("Profitable All 4", r))

    # Best in World players not at $1,500+
    bitw_low = [r for r in results if r["tier"] == "Best in the World" and "$1,500" not in r["rec"]]
    for r in bitw_low[:30]:
        edge_cases.append(("BITW not $1.5K", r))

    # High FT rate relative to CR (deep runners)
    deep_runners = [r for r in results if r["ft_rate"] > r["cash_rate"] * 0.55 and r["cash_rate"] > 0.15]
    deep_runners.sort(key=lambda r: -r["ft_rate"] / max(0.01, r["cash_rate"]))
    for r in deep_runners[:30]:
        edge_cases.append(("Deep Runner", r))

    for row_idx, (label, r) in enumerate(edge_cases[:200], 2):
        vals = [label, r["tier"], r["tournaments"], r["cash_rate"], r["ft_rate"],
                r["top3_conv"], r["ev_low"], r["ev_mid_low"], r["ev_mid"], r["ev_high"], r["rec"]]
        for col_idx, val in enumerate(vals, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            if col_idx in (4, 5, 6):
                cell.number_format = "0.0%"
            elif col_idx in (7, 8, 9, 10):
                cell.number_format = '$#,##0'
            if col_idx == 11:
                color = rec_color(val)
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(s4_headers) + 1):
        ws4.column_dimensions[get_column_letter(col)].width = 16
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions[get_column_letter(len(s4_headers))].width = 28
    ws4.freeze_panes = "B2"

    return wb


def main():
    start = time.time()

    print("Loading 100K synthetic players...")
    players = load_synthetic_players("c:/Users/nicho/OneDrive/Desktop/code/Poker_videos/synthetic_players_100k.csv")
    print(f"Loaded {len(players):,} players in {time.time()-start:.1f}s\n")

    print("Running 4-tier stake move model...")
    t0 = time.time()
    results = [r for r in (evaluate_synthetic(p) for p in players) if r]
    print(f"Evaluated {len(results):,} players in {time.time()-t0:.1f}s\n")

    # Gather all recs
    all_recs = sorted(set(r["rec"] for r in results),
                      key=lambda x: (0 if "DOWN" in x else 1 if "STAY" in x else 2 if "CONSIDER" in x else 3))

    # Print by player tier
    print(f"{'Player Tier':>22s} | {'Count':>6s}", end="")
    for rec in all_recs:
        short = rec.replace("MOVE UP to ", "UP ").replace("CONSIDER ", "CON ").replace("STAY at ", "STAY ")
        print(f" | {short:>12s}", end="")
    print(f" | {'%MoveUp':>7s}")
    print("-" * (35 + 15 * len(all_recs) + 10))

    for tn in PLAYER_TIERS:
        tr = [r for r in results if r["tier"] == tn]
        counts = Counter(r["rec"] for r in tr)
        n = len(tr)
        mu = sum(v for k, v in counts.items() if "MOVE UP" in k) / max(1, n)
        print(f"{tn:>22s} | {n:>6,d}", end="")
        for rec in all_recs:
            print(f" | {counts.get(rec, 0):>12,d}", end="")
        print(f" | {mu:>6.1%}")

    # Overall
    total = Counter(r["rec"] for r in results)
    print(f"\n{'TOTAL':>22s}: {len(results):,}")
    for rec in all_recs:
        pct = total.get(rec, 0) / len(results) * 100
        print(f"  {rec:>35s}: {total.get(rec,0):>6,d} ({pct:.1f}%)")

    # Boundary CRs
    print("\n--- Decision Boundary Cash Rates ---")
    for rec in all_recs:
        crs = [r["cash_rate"] for r in results if r["rec"] == rec]
        if crs:
            print(f"  {rec:>35s}: P10={np.percentile(crs,10):.1%}  median={np.median(crs):.1%}  P90={np.percentile(crs,90):.1%}")

    # Edge cases
    print("\n--- Edge Cases ---")
    losing_low = sum(1 for r in results if r["ev_low"] <= 0 and "MOVE UP" in r["rec"])
    all_profit = sum(1 for r in results if r["ev_low"] > 0 and r["ev_mid_low"] > 0 and r["ev_mid"] > 0 and r["ev_high"] > 0)
    small_hero = sum(1 for r in results if "MOVE UP" in r["rec"] and r["tournaments"] < 80)
    print(f"  Losing at $300 but MOVE UP higher: {losing_low:,}")
    print(f"  Profitable at ALL 4 tiers: {all_profit:,}")
    print(f"  Small sample (<80) MOVE UP: {small_hero:,}")

    # Save
    print("\nBuilding Excel...")
    wb = build_workbook(results)
    output_path = "c:/Users/nicho/OneDrive/Desktop/code/Poker_videos/bulk_stake_analysis_100k.xlsx"
    wb.save(output_path)
    print(f"Saved -> {output_path}")
    print(f"\nTotal time: {time.time()-start:.1f}s")


if __name__ == "__main__":
    main()
