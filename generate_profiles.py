"""
Generate 100K synthetic player profiles with realistic statistical variation.

Uses normal distributions with meaningful variance so that:
- Stats within a tier overlap with adjacent tiers
- Outliers emerge naturally (high CR + low FT, small sample heroes, etc.)
- Cash rate, FT rate, and top-3 conversion are partially independent
  (correlated but not locked together like real players)

This feeds into stake_move_model.py for bulk recommendation testing.
"""

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

np.random.seed(42)

NUM_PLAYERS = 100_000

# ============================================================
# TIER DEFINITIONS
# ============================================================
# Each tier defines the CENTER of the distribution (mean) and spread (std).
# Stats are drawn from normal distributions so tiers overlap naturally.
#
# Format: (tier_name, count, tournament_params, cash_rate_params,
#           ft_rate_params, top3_conv_params)
# Each param is (mean, std, min_clip, max_clip)

# Distribution calibrated to reality: ~80% of poker players are losing.
# Bottom tiers have lower cash rates than before — a true recreational
# player cashes 6-8%, barely above random chance in a 12.5% payout structure.
# The ones cashing 10%+ already have SOME edge.
#
# Population distribution:
#   Bottom 55%: Brand New + Bad Reg (the 87.5% who lose every tournament)
#   Next 20%: Average + Slightly Profitable (break-even zone)
#   Next 15%: Good Reg + Great Reg (consistent small winners)
#   Top 10%: Low Pro through Best in World (the sharks)
#   Top 1%: Mid Pro and above

TIERS = [
    {
        "name": "Brand New",
        "count": 30000,  # 30% — huge population, most quit quickly
        "tournaments": (15, 12, 1, 60),  # 1 every 2 weeks, most quit within a year or two
        "cash_rate":   (0.06, 0.03, 0.00, 0.14),  # barely above random
        "ft_rate":     (0.010, 0.010, 0.00, 0.04),  # almost never reach FT
        "top3_conv":   (0.15, 0.12, 0.00, 0.50),  # lucky when they do
    },
    {
        "name": "Bad Reg",
        "count": 29000,  # 29% — the guys who keep coming back but don't improve
        "tournaments": (80, 60, 15, 300),  # 1-2/week for a year or two
        "cash_rate":   (0.09, 0.03, 0.03, 0.16),  # below 12.5% = losing to rake
        "ft_rate":     (0.025, 0.015, 0.00, 0.06),
        "top3_conv":   (0.20, 0.10, 0.00, 0.45),
    },
    {
        "name": "Average Player",
        "count": 18000,  # 18% — the break-even zone, might be +/- slightly
        "tournaments": (150, 100, 30, 500),  # 1-2/week, been playing a few years
        "cash_rate":   (0.13, 0.03, 0.07, 0.20),  # right around the 12.5% payout line
        "ft_rate":     (0.04, 0.02, 0.01, 0.09),
        "top3_conv":   (0.28, 0.10, 0.05, 0.55),
    },
    {
        "name": "Slightly Profitable",
        "count": 10000,  # 10% — first tier that's clearly winning
        "tournaments": (250, 150, 50, 800),  # 2-3/week, committed for a few years
        "cash_rate":   (0.17, 0.03, 0.10, 0.25),
        "ft_rate":     (0.06, 0.025, 0.02, 0.12),
        "top3_conv":   (0.33, 0.09, 0.10, 0.58),
    },
    {
        "name": "Good Reg",
        "count": 7000,  # 7% — solid winners
        "tournaments": (400, 200, 100, 1200),  # 3-4/week, been at it for years
        "cash_rate":   (0.22, 0.03, 0.15, 0.30),
        "ft_rate":     (0.09, 0.03, 0.03, 0.16),
        "top3_conv":   (0.40, 0.08, 0.18, 0.65),
    },
    {
        "name": "Great Reg",
        "count": 4000,  # 4% — the guys who grind a real living
        "tournaments": (600, 300, 150, 1800),  # 4-5/week, serious grinder
        "cash_rate":   (0.27, 0.03, 0.20, 0.36),
        "ft_rate":     (0.12, 0.03, 0.05, 0.20),
        "top3_conv":   (0.45, 0.08, 0.25, 0.68),
    },
    {
        "name": "Low Level Pro",
        "count": 1200,  # 1.2% — grinders making a modest living
        "tournaments": (1200, 500, 300, 3500),
        "cash_rate":   (0.30, 0.03, 0.23, 0.40),
        "ft_rate":     (0.15, 0.03, 0.07, 0.24),
        "top3_conv":   (0.48, 0.07, 0.28, 0.70),
    },
    {
        "name": "Mid Level Pro",
        "count": 550,  # 0.55% — solid full-time pros
        "tournaments": (1800, 800, 400, 5000),
        "cash_rate":   (0.33, 0.025, 0.26, 0.42),
        "ft_rate":     (0.18, 0.03, 0.10, 0.26),
        "top3_conv":   (0.52, 0.06, 0.33, 0.73),
    },
    {
        "name": "High Level Pro",
        "count": 200,  # 0.2% — the names you recognize
        "tournaments": (2500, 1200, 500, 8000),
        "cash_rate":   (0.36, 0.025, 0.30, 0.46),
        "ft_rate":     (0.20, 0.03, 0.12, 0.30),
        "top3_conv":   (0.55, 0.06, 0.36, 0.76),
    },
    {
        "name": "Best in the World",
        "count": 50,  # 0.05% — the absolute elite
        "tournaments": (4000, 2000, 800, 12000),
        "cash_rate":   (0.40, 0.03, 0.33, 0.52),
        "ft_rate":     (0.24, 0.035, 0.16, 0.36),
        "top3_conv":   (0.60, 0.06, 0.40, 0.80),
    },
]


def draw(mean, std, lo, hi):
    """Draw from a clipped normal distribution."""
    return float(np.clip(np.random.normal(mean, std), lo, hi))


def generate_player(tier, player_id):
    """Generate a single player profile with realistic variance.

    Aggression is drawn independently and then SHAPES the stats:
    - Aggressive players: lower CR, higher FT rate per cash, more 1sts
    - Passive players: higher CR, lower FT rate per cash, more min-cashes

    This means two players with the same underlying skill (tier) can
    have very different stat lines based on playstyle.
    """
    t_mean, t_std, t_lo, t_hi = tier["tournaments"]
    tournaments = max(int(t_lo), int(np.clip(np.random.normal(t_mean, t_std), t_lo, t_hi)))

    # Draw aggression independently — normally distributed around 0.50
    # with meaningful spread so we get nits (0.15) and maniacs (0.85)
    aggression = float(np.clip(np.random.normal(0.50, 0.18), 0.05, 0.95))
    agg_delta = aggression - 0.5  # -0.45 to +0.45

    # Base cash rate from tier, then adjust for aggression
    # Aggressive players cash LESS (trade CR for stack size)
    base_cr = draw(*tier["cash_rate"])
    cr_adj = 1.0 - agg_delta * 0.5  # aggressive: 0.75x CR, passive: 1.25x CR
    cash_rate = max(0.02, min(0.55, base_cr * cr_adj))

    # FT rate: correlated with cash rate + skill, but aggression boosts it
    # Aggressive players get to FT more often per cash (bigger stacks)
    cr_mean = tier["cash_rate"][0]
    ft_mean, ft_std, ft_lo, ft_hi = tier["ft_rate"]
    cr_deviation = (base_cr - cr_mean) / max(0.01, tier["cash_rate"][1])
    ft_adjusted_mean = ft_mean + cr_deviation * ft_std * 0.4

    # Aggression boosts FT rate: big stacks go deeper
    stack_mult = 8.0 * (1.0 + agg_delta * 1.5)
    ft_agg_boost = (stack_mult / 8.0) ** 0.7  # same as apply_aggression
    ft_rate = float(np.clip(np.random.normal(ft_adjusted_mean * ft_agg_boost, ft_std), ft_lo, ft_hi))
    ft_rate = min(ft_rate, cash_rate * 0.85)

    # Top-3 conversion: aggression helps (big stacks dominate FTs)
    base_t3 = draw(*tier["top3_conv"])
    t3_boost = (stack_mult / 8.0) ** 0.5
    top3_conv = min(0.85, base_t3 * t3_boost)

    # Build placement distribution
    total_cashes = max(0, round(tournaments * cash_rate))
    total_fts = max(0, round(tournaments * ft_rate))
    total_fts = min(total_fts, total_cashes)

    top3_total = max(0, round(total_fts * top3_conv))
    top3_total = min(top3_total, total_fts)

    # Top-3 split: aggression skews toward 1st place (big stacks win more)
    if top3_total == 0:
        first, second, third = 0, 0, 0
    else:
        # Base alphas skew toward 3rd (hardest to win), aggression shifts to 1st
        a1 = 0.8 + agg_delta * 0.6   # aggressive: 1.1, passive: 0.5
        a2 = 1.0
        a3 = 1.2 - agg_delta * 0.4   # aggressive: 1.0, passive: 1.4
        weights = np.random.dirichlet([max(0.2, a1), max(0.2, a2), max(0.2, a3)])
        first = max(0, round(top3_total * weights[0]))
        second = max(0, round(top3_total * weights[1]))
        third = max(0, top3_total - first - second)

    fourth_ninth = max(0, total_fts - first - second - third)
    tenth_plus = max(0, total_cashes - first - second - third - fourth_ninth)

    # Recalculate actual rates from integer counts
    actual_cashes = first + second + third + fourth_ninth + tenth_plus
    actual_cr = actual_cashes / tournaments if tournaments > 0 else 0
    actual_ft = (first + second + third + fourth_ninth) / tournaments if tournaments > 0 else 0
    actual_top3_conv = (first + second + third) / max(1, first + second + third + fourth_ninth)

    return {
        "id": player_id,
        "tier": tier["name"],
        "tournaments": tournaments,
        "1st": first,
        "2nd": second,
        "3rd": third,
        "4th-9th": fourth_ninth,
        "10th+": tenth_plus,
        "total_cashes": actual_cashes,
        "cash_rate": round(actual_cr, 4),
        "ft_rate": round(actual_ft, 4),
        "top3_conv": round(actual_top3_conv, 4),
        "aggression": round(aggression, 3),
    }


def generate_all():
    """Generate all players across all tiers."""
    players = []
    player_id = 1
    for tier in TIERS:
        for _ in range(tier["count"]):
            players.append(generate_player(tier, player_id))
            player_id += 1
    return players


def save_csv(players, path):
    """Save to CSV for fast loading by the stake move model."""
    import csv
    fields = ["id", "tier", "tournaments", "1st", "2nd", "3rd", "4th-9th",
              "10th+", "total_cashes", "cash_rate", "ft_rate", "top3_conv", "aggression"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(players)


def save_xlsx_summary(players, path):
    """Save a summary Excel with tier distributions (not all 100K rows)."""
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    tier_colors = {
        "Brand New": "FFE0E0", "Bad Reg": "FFD0C0", "Average Player": "FFE8B0",
        "Slightly Profitable": "FFFFC0", "Good Reg": "D8F0D8", "Great Reg": "B0E0B0",
        "Low Level Pro": "B0D8F0", "Mid Level Pro": "90C0E8",
        "High Level Pro": "C0B0F0", "Best in the World": "FFD700",
    }

    # Sheet 1: Tier Summary
    ws1 = wb.active
    ws1.title = "Tier Summary"

    headers = ["Tier", "Count", "Avg Tournaments", "Avg Cash Rate", "Std CR",
               "Avg FT Rate", "Std FT", "Avg Top3 Conv",
               "Avg 1st/t", "Avg 2nd/t", "Avg 3rd/t",
               "Min CR", "P10 CR", "P25 CR", "Median CR", "P75 CR", "P90 CR", "Max CR"]

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    tier_names = [t["name"] for t in TIERS]
    for row_idx, tier_name in enumerate(tier_names, 2):
        tp = [p for p in players if p["tier"] == tier_name]
        crs = [p["cash_rate"] for p in tp]
        fts = [p["ft_rate"] for p in tp]
        t3s = [p["top3_conv"] for p in tp]
        tourns = [p["tournaments"] for p in tp]

        fill = PatternFill(start_color=tier_colors.get(tier_name, "FFFFFF"),
                          end_color=tier_colors.get(tier_name, "FFFFFF"), fill_type="solid")

        vals = [
            tier_name, len(tp), round(np.mean(tourns)),
            round(np.mean(crs), 4), round(np.std(crs), 4),
            round(np.mean(fts), 4), round(np.std(fts), 4),
            round(np.mean(t3s), 4),
            round(np.mean([p["1st"] / p["tournaments"] for p in tp]), 5),
            round(np.mean([p["2nd"] / p["tournaments"] for p in tp]), 5),
            round(np.mean([p["3rd"] / p["tournaments"] for p in tp]), 5),
            round(min(crs), 4),
            round(np.percentile(crs, 10), 4),
            round(np.percentile(crs, 25), 4),
            round(np.percentile(crs, 50), 4),
            round(np.percentile(crs, 75), 4),
            round(np.percentile(crs, 90), 4),
            round(max(crs), 4),
        ]

        for col_idx, val in enumerate(vals, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if col_idx in (4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18):
                cell.number_format = "0.0%"

    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 16
    ws1.column_dimensions["A"].width = 22
    ws1.freeze_panes = "B2"

    # Sheet 2: Overlap analysis — how many players in each tier could be confused for another
    ws2 = wb.create_sheet("Tier Overlap")
    ws2.cell(row=1, column=1, value="Cash Rate Distribution Overlap Between Tiers").font = Font(bold=True, size=12)

    ws2.cell(row=3, column=1, value="Tier").font = Font(bold=True)
    for col_idx, tn in enumerate(tier_names, 2):
        ws2.cell(row=3, column=col_idx, value=tn).font = Font(bold=True, size=9)

    for row_idx, tn1 in enumerate(tier_names, 4):
        ws2.cell(row=row_idx, column=1, value=tn1).font = Font(bold=True, size=9)
        crs1 = [p["cash_rate"] for p in players if p["tier"] == tn1]
        r1_lo, r1_hi = np.percentile(crs1, 10), np.percentile(crs1, 90)

        for col_idx, tn2 in enumerate(tier_names, 2):
            crs2 = [p["cash_rate"] for p in players if p["tier"] == tn2]
            # What % of tier2 falls within tier1's 10-90 range?
            overlap = np.mean([(r1_lo <= cr <= r1_hi) for cr in crs2])
            cell = ws2.cell(row=row_idx, column=col_idx, value=round(overlap, 3))
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="center")
            # Color intensity by overlap
            if overlap > 0.5:
                cell.fill = PatternFill(start_color="ff6b6b", end_color="ff6b6b", fill_type="solid")
            elif overlap > 0.2:
                cell.fill = PatternFill(start_color="ffd93d", end_color="ffd93d", fill_type="solid")
            elif overlap > 0.05:
                cell.fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")

    for col in range(1, len(tier_names) + 2):
        ws2.column_dimensions[get_column_letter(col)].width = 16

    # Sheet 3: Sample of 200 players (for spot-checking)
    ws3 = wb.create_sheet("Sample (200 players)")
    sample_headers = ["ID", "Tier", "Tournaments", "1st", "2nd", "3rd", "4th-9th",
                      "10th+", "Total Cashes", "Cash Rate", "FT Rate", "Top3 Conv"]

    for col, h in enumerate(sample_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Sample 20 from each tier
    sample = []
    for tn in tier_names:
        tier_players = [p for p in players if p["tier"] == tn]
        indices = np.random.choice(len(tier_players), size=min(20, len(tier_players)), replace=False)
        sample.extend([tier_players[i] for i in indices])

    for row_idx, p in enumerate(sample, 2):
        fill = PatternFill(start_color=tier_colors.get(p["tier"], "FFFFFF"),
                          end_color=tier_colors.get(p["tier"], "FFFFFF"), fill_type="solid")
        vals = [p["id"], p["tier"], p["tournaments"], p["1st"], p["2nd"], p["3rd"],
                p["4th-9th"], p["10th+"], p["total_cashes"], p["cash_rate"],
                p["ft_rate"], p["top3_conv"]]
        for col_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if col_idx in (10, 11, 12):
                cell.number_format = "0.0%"

    for col in range(1, len(sample_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 15
    ws3.freeze_panes = "C2"

    wb.save(path)


def main():
    start = time.time()
    print(f"Generating {NUM_PLAYERS:,} player profiles...\n")

    players = generate_all()
    gen_time = time.time() - start
    print(f"Generated in {gen_time:.1f}s\n")

    # Print tier summary
    print(f"{'Tier':>22s} | {'Count':>6s} | {'Avg Tourn':>9s} | {'Avg CR':>7s} | {'Std CR':>7s} | {'P10':>6s} | {'P90':>6s} | {'Avg FT':>7s} | {'Avg T3C':>7s}")
    print("-" * 105)
    for tier in TIERS:
        tp = [p for p in players if p["tier"] == tier["name"]]
        crs = [p["cash_rate"] for p in tp]
        fts = [p["ft_rate"] for p in tp]
        t3s = [p["top3_conv"] for p in tp]
        tourns = [p["tournaments"] for p in tp]
        print(f"{tier['name']:>22s} | {len(tp):>6,d} | {np.mean(tourns):>9.0f} | {np.mean(crs):>6.1%} | {np.std(crs):>6.1%} | {np.percentile(crs,10):>5.1%} | {np.percentile(crs,90):>5.1%} | {np.mean(fts):>6.1%} | {np.mean(t3s):>6.1%}")

    # Save CSV (fast, for model consumption)
    csv_path = "c:/Users/nicho/OneDrive/Desktop/code/Poker_videos/synthetic_players_100k.csv"
    save_csv(players, csv_path)
    print(f"\nCSV saved -> {csv_path}")

    # Save Excel summary
    xlsx_path = "c:/Users/nicho/OneDrive/Desktop/code/Poker_videos/player_profiles_100k_summary.xlsx"
    save_xlsx_summary(players, xlsx_path)
    print(f"Excel summary saved -> {xlsx_path}")

    total_time = time.time() - start
    print(f"\nTotal time: {total_time:.1f}s")


if __name__ == "__main__":
    main()
